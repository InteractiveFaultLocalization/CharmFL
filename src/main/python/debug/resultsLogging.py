import openpyxl

def results_on_file(resultsLoggingPath, project, bug, rlist, slist, statslist):

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("result")

    count = 0
    for r, s, st in zip(rlist, slist, statslist):
        for i in range(len(r)):
            ws.cell(row=i+1, column=1+count).value = r[i]
            ws.cell(row=i+1, column=2+count).value = s[i]
            ws.cell(row=i+1, column=3+count).value = st[i]   
        count = count + 3

    wb.save(resultsLoggingPath + "/bugs/" + project + "_bug" + bug + ".xlsx")
