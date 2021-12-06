import pytest
import openpyxl

def get_tests_results(path):
    #path = "C://Users//user//Documents//RajmondFL//CharmFL//example//"
    pytest.main(["-v", "--excelreport=" + path + "report.xlsx", path])
    test_names = []
    complete_test_names = []
    test_results = []
    results_dict = {}

    wb_obj = openpyxl.load_workbook(path + "report.xlsx")
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row - 1

    for i in range(m_row):
        t_n = sheet_obj.cell(row = i + 2, column = 2).value
        c_t_n = sheet_obj.cell(row = i + 2, column = 1).value + "." + t_n
        t_r = sheet_obj.cell(row = i + 2, column = 4).value

        if t_r == "PASSED" or t_r == "FAILED":
            test_names.append(t_n)
            complete_test_names.append(c_t_n)
            test_results.append(t_r)
            results_dict[c_t_n] = t_r

    if results_dict:
        print(test_names)
        print(complete_test_names)
        print(test_results)
        print(results_dict)
        return results_dict
    else:
        raise Exception('No test results collected!')