import os
import sys
import pytest
import openpyxl

from constans import EXCEL_REPORT_FILE_NAME
from error_codes import TEST_FOLDERS_NOT_FOUND, EXCEL_REPORT_FILE_NOT_FOUND, RESULT_DICT_EMPTY


def get_test_files(project_path):
    test_folder = ""
    tests_folder = ""
    test_files = []
    tests_files = []

    if os.path.isdir(project_path + os.path.sep + "tests"):
        tests_folder = "tests"
    if os.path.isdir(project_path + os.path.sep + "test"):
        test_folder = "test"

    if tests_folder == "" and test_folder == "":
        sys.exit(TEST_FOLDERS_NOT_FOUND)

    if tests_folder != "":
        with os.scandir(project_path + os.path.sep + tests_folder) as it:
            for entry in it:
                if not entry.name.startswith('.') and not entry.name.startswith('..') and not entry.name == "__init__.py" and entry.is_file() and entry.name.endswith('.py'):
                    tests_files.append(entry.name)

    if test_folder != "":
        with os.scandir(project_path + os.path.sep + test_folder) as it:
            for entry in it:
                if not entry.name.startswith('.') and not entry.name.startswith('..') and not entry.name == "__init__.py" and entry.is_file() and entry.name.endswith(".py"):
                    test_files.append(entry.name)

    return test_folder, tests_folder, test_files, tests_files


def get_tests_results(project_path):
    test_folder, tests_folder, test_files, tests_files = get_test_files(project_path)

    pytest_args = ["-v"]
    for file in tests_files:
        pytest_args.append(project_path + os.path.sep + tests_folder + os.path.sep + file)
    for file in test_files:
        pytest_args.append(project_path + os.path.sep + test_folder + os.path.sep + file)

    pytest_args.append("--excelreport=" + project_path + os.path.sep + EXCEL_REPORT_FILE_NAME)


    pytest.main(pytest_args)
    report_file = project_path + os.path.sep + EXCEL_REPORT_FILE_NAME

    if not os.path.exists(report_file):
        sys.exit(EXCEL_REPORT_FILE_NOT_FOUND)

    wb_obj = openpyxl.load_workbook(EXCEL_REPORT_FILE_NAME)
    sheet_obj = wb_obj.active

    results_dict = {}

    for i in range(sheet_obj.max_row - 1):
        suite_name = sheet_obj.cell(row=i + 2, column=1).value
        test_name = sheet_obj.cell(row=i + 2, column=2).value
        test_result = sheet_obj.cell(row=i + 2, column=4).value
        test_file_name = sheet_obj.cell(row=i + 2, column=8).value
        test_file_name = test_file_name.replace("/", ".").replace("\\", ".").replace(".py", "")

        if test_file_name == suite_name:
            coverage_test_name = test_file_name + "." + test_name
        else:
            coverage_test_name = test_file_name + "." + suite_name + "." + test_name

        if test_result == "PASSED" or test_result == "FAILED":
            results_dict[coverage_test_name] = test_result

    if results_dict:
        return results_dict
    else:
        sys.exit(RESULT_DICT_EMPTY)
