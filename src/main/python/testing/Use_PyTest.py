import os
import sys
import pytest
import openpyxl
import json

from constans import EXCEL_REPORT_FILE_NAME, JSON_REPORT_FILE_NAME
from error_codes import EXCEL_REPORT_FILE_NOT_FOUND, RESULT_DICT_EMPTY, JSON_REPORT_FILE_NOT_FOUND
from testing.Test_Utils import Test_Utils


class Use_Pytest(Test_Utils):
    project_path = ""
    results_dict = {}
    number_of_failed_tests = 0
    number_of_passed_tests = 0

    def __init__(self, project_path):
        self.project_path = project_path

    def run_tests(self):
        test_folder, tests_folder, test_files, tests_files = super(Use_Pytest, self).get_test_files(self.project_path)

        pytest_args = ["-v", "--cov=.", "--cov-context=test", "--json-report"]
        for file in tests_files:
            pytest_args.append(self.project_path + os.path.sep + tests_folder + os.path.sep + file)
        for file in test_files:
            pytest_args.append(self.project_path + os.path.sep + test_folder + os.path.sep + file)

        # pytest_args.append("--excelreport=" + self.project_path + os.path.sep + EXCEL_REPORT_FILE_NAME) ### Deprecated, we wont use this hopefully. Should be deleted when the json is tested thoroughly.

        pytest.main(pytest_args)
        self.__get_test_results_from_json(self.results_dict)
        # self.__make_excel_report(self.results_dict) ### Deprecated, we wont use this hopefully. Should be deleted when the json is tested thoroughly.

    def get_tests_results(self):

        if self.results_dict:
            return self.results_dict
        else:
            sys.exit(RESULT_DICT_EMPTY)

    ### Deprecated, we wont use this hopefully. Should be deleted when the json is tested thoroughly.
    def __make_excel_report(self, results_dict):
        report_file = self.project_path + os.path.sep + EXCEL_REPORT_FILE_NAME

        if not os.path.exists(report_file):
            sys.exit(EXCEL_REPORT_FILE_NOT_FOUND)

        wb_obj = openpyxl.load_workbook(EXCEL_REPORT_FILE_NAME)
        sheet_obj = wb_obj.active

        for i in range(sheet_obj.max_row - 1):
            suite_name = sheet_obj.cell(row=i + 2, column=1).value
            test_name = sheet_obj.cell(row=i + 2, column=2).value
            test_result = sheet_obj.cell(row=i + 2, column=4).value
            test_file_name = sheet_obj.cell(row=i + 2, column=8).value
            test_file_name = test_file_name.replace("\\","/")
            coverage_test_name = test_file_name + "::" + test_name+ "|run"

            if test_result == "PASSED":
                self.number_of_passed_tests = self.number_of_passed_tests + 1
            if test_result == "FAILED":
                self.number_of_failed_tests = self.number_of_failed_tests + 1
            if test_result == "PASSED" or test_result == "FAILED":
                results_dict[coverage_test_name] = test_result

    def __get_test_results_from_json(self, results_dict):
        report_file_path = self.project_path + os.path.sep + JSON_REPORT_FILE_NAME

        if not os.path.exists(report_file_path):
            sys.exit(JSON_REPORT_FILE_NOT_FOUND)
        report_json = open(report_file_path)
        data = json.load(report_json)
        report_json.close()
        for test_object in data["tests"]:
            coverage_test_name = test_object["nodeid"] + "|run"
            test_result =str(test_object["outcome"]).upper()
            if test_result == "PASSED":
                self.number_of_passed_tests = self.number_of_passed_tests + 1
            if test_result == "FAILED":
                self.number_of_failed_tests = self.number_of_failed_tests + 1
            if test_result == "PASSED" or test_result == "FAILED":
                results_dict[coverage_test_name] = test_result


    def get_number_of_failed_test_cases(self):
        return self.number_of_failed_tests

    def get_number_of_passed_test_cases(self):
        return self.number_of_passed_tests